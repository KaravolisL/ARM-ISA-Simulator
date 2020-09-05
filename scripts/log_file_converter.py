# Standard Libraries
from argparse import ArgumentParser
from openpyxl import Workbook, worksheet
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
import os
import sys

import logging
logger = logging.getLogger(__name__)

EXIT_SUCCESS = 0

PRINT_GROUP_FILLS = {
    "GENERAL" : PatternFill(start_color='FABF8F', fill_type='solid'),
    "MEMORY" : PatternFill(start_color='92CDDC', fill_type='solid'),
    "INSTRUCTION" : PatternFill(start_color='C4D79B', fill_type='solid'),
    "ASSERT" : PatternFill(start_color='C0504D', fill_type='solid')
}

def stylize(sheet):
    """Styles each cell

    :param worksheet sheet: Sheet to modify
    
    """
    column_widths = []
    for row in sheet.iter_rows():
        # Determine the background color for this row
        try:
            background_color = PRINT_GROUP_FILLS[row[1].value]
        except KeyError:
            if row[1].value != "Print Group":
                logger.info("Background color not found for group %s. Consider adding one!", row[1].value)
            background_color = PatternFill(start_color='FFFFFF', fill_type='solid')
        
        for i, cell in enumerate(row):
            cell.fill = background_color
            try:
                column_widths[i] = max(column_widths[i], len(str(cell.value)))
            except IndexError:
                column_widths.append(len(str(cell.value)))

    for i, column_width in enumerate(column_widths):
        # Add a little extra width with the 2
        sheet.column_dimensions[get_column_letter(i + 1)].width = column_width + 2

def convert_log_file(input_file_name, output_file_name):
    """Converts a log file outputted by the program to an excel file

    :param str input_file_name: Name of the file to convert
    :param str output_file_name: Name of the file that will be outputted

    """
    logger.info('Input file name: %s', input_file_name)
    logger.info('Output file name: %s', output_file_name)

    workbook = Workbook()
    sheet = workbook.active

    # Print the headers in the first row
    headers = ('Level', 'Print Group', 'File', 'Function', 'Line', 'Message')
    for i, header in enumerate(headers):
        sheet.cell(row=1, column=i+1).value = header
        sheet.cell(row=1, column=i+1).font = Font(bold=True)

    # Convert the lines in the debug file
    with open(input_file_name, 'r') as input_file:
        for line in input_file:
            entries = line.split(' : ', len(headers) - 1)
            if len(entries) < len(headers):
                logger.warning("Invalid formatting detected on line %s", line)
            
            # Print the lines to the cells
            sheet.append(entries)
    logger.info("Dimensions of sheet: %s", sheet.dimensions)

    # Apply style to cells
    stylize(sheet)

    # Add filters to the headers
    sheet.auto_filter.ref = sheet.dimensions

    workbook.save(filename=output_file_name)

def main():
    argument_parser = ArgumentParser(
        prog='python log_file_formatter.py log_file',
        description='Converts a log file to a formatted excel file'
    )
    argument_parser.add_argument('log_file')
    argument_parser.add_argument('--in_place', action='store_true', help='Saves the file in the location of this script')
    argument_parser.add_argument('--verbose', '-v', action='store_true')
    args = argument_parser.parse_args()

    if not os.path.isfile(args.log_file):
        raise ValueError("File %s does not exist", args.memory_file)

    if args.verbose:
        logging.basicConfig(level=logging.DEBUG)
    else:
        logging.basicConfig(level=logging.INFO)

    if os.path.splitext(args.log_file)[1] != '.log':
        logger.warning("A file with extension .log is expected. Resulting file may be incorrect")

    if args.in_place:
        output_file_name = os.path.splitext(os.path.basename(args.log_file))[0] + '.xlsx'
    else:
        output_file_name = os.path.splitext(args.log_file)[0] + '.xlsx'
    logger.info("Absolute path of output file here: %s", os.path.abspath(output_file_name))

    convert_log_file(args.log_file, output_file_name)

    logger.info("Log file conversion complete: SUCCESS")

    return EXIT_SUCCESS


if __name__ == '__main__':
    sys.exit(main())
